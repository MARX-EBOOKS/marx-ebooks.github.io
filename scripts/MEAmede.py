import re
import uuid
from datetime import datetime
from pathlib import Path
import shutil
from html.parser import HTMLParser
from urllib.parse import urlparse, urljoin
import base64
from bs4 import BeautifulSoup
import openpyxl
import chapternumbergerman
import json
class HeadingExtractor:
    """HTML标题层级提取器 (h1-h6)，使用BeautifulSoup"""
    def __init__(self):
        self.headings = []
    
    def extract_headings_from_html(self, html_content, source_file, chapter_number,title="",volume_number=0):
        """从HTML内容中提取所有标题层级并添加ID"""
        try:
            soup = BeautifulSoup(html_content,'html.parser')
            headings_data = []
            
            # 查找所有标题标签
            heading_tags = soup.find_all(['h1','h2', 'h3', 'h4', 'h5', 'h6'])
            for i, tag in enumerate(heading_tags):
                text = tag.get_text(strip=True,separator='  ')
                if text and len(text) > 0:
                    # 为标题添加锚点ID                    
                    text=text.replace("<","&lt;")
                    text=text.replace(">","&gt;")
                    text=re.sub(r'￥￥￥[\S ]+?￥￥￥',r'',text,flags=re.DOTALL|re.IGNORECASE)
                    search_words=[r'Karl Marx',r'Friedrich Engels',r'Lesart',r'Fußnote']
                    if any(word in text for word in search_words) and not "von Karl Marx" in text:
                        continue      
                    if i==0 and not source_file.startswith("2000"):
                        title=text                                      
                    # 确定标题级别
                    if volume_number in range(23,26):
                        text=text.replace(" Abschnitt "," Abschnitt. ").replace(" Kapitel "," Kapitel. ")
                    anchor_id = f"h{chapter_number}-{i+1}"
                    tag['id']=anchor_id                    
                    #tag['id']=anchor_id
                    level = int(tag.name[1])  # h1->1, h2->2, etc.
                    headings_data.append({
                        'tag': tag.name,
                        'text': text,
                        'level': level,
                        'id': anchor_id,
                        'source_file': source_file,
                        'chapter_number': chapter_number
                    })         
            #check_tags = soup.find_all(['h1','h2', 'h3', 'h4', 'h5', 'h6','b','i','a'])
            #hazard=[]
            

            #for tag in check_tags:
             #   if len(tag.text)>=150:
              #      print(f"{source_file}中有异常：{tag.text}")
              #      hazard.append(f'<{tag.name}>{tag.text}')
            #if hazard:
            #    return f'共{len(hazard)}处异常文本:{chr(10).join(hazard)}', headings_data
            fixed_content=str(soup.body)
            if not soup.body:
                #print(f"{source_file}缺失body\n")
                fixed_content=re.sub(r'</title>','</title></head>',str(soup),flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'</title>(</head>)*[\s\r\n\S]*</head>',f'</title>\n</head>\n<body>',fixed_content,flags=re.DOTALL | re.IGNORECASE)
                soup2 = BeautifulSoup(fixed_content, 'html.parser')
                fixed_content=str(soup2.body)
                if not fixed_content or re.match(r'<body>[\s\r\n]*</body>', fixed_content,flags=re.DOTALL | re.IGNORECASE):
                    print(f"{source_file}仍缺失body\n")
            fixed_content=re.sub(r'<body[ \S]*?>',r'<body>',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
            final_content=f"""<html lang="de">
<head>
<META name="viewport" content="width=device-width, initial-scale=1.0"/>
<META content="text/html; charset=UTF-8" http-equiv="Content-Type"/>
    <title>{title}</title>
<link rel="stylesheet" type="text/css" href="../mewde.css"/>
<script src="/mlr.js"></script>
</head>
{fixed_content}
</html>"""
            return final_content,headings_data
            
        except Exception as e:
            print(f"  警告: 解析标题时出错: {e}")
            return html_content, []
class TitleExtractor(HTMLParser):
    """HTML标题提取器"""
    def __init__(self):
        super().__init__()
        self.title = ""
        self.in_title = False
        self.all_headings = []
    
    def handle_starttag(self, tag, attrs):
        if tag.lower() == 'title':
            self.in_title = True
    
    def handle_endtag(self, tag):
        if tag.lower() == 'title':
            self.in_title = False
    
    def handle_data(self, data):
        if self.in_title:
            data=re.sub(r'\[\d+?\]','',data)
            data=re.sub(r'\[注：[\s\S]+?\]','',data)
            self.title += data


class LinkExtractor(HTMLParser):
    """HTML链接提取器"""
    def __init__(self):
        super().__init__()
        self.links = []
        self.current_link = None
        self.in_link = False
        self.link_text = ""
    
    def handle_starttag(self, tag, attrs):
        if tag.lower() == 'a':
            for name, value in attrs:
                if name.lower() == 'href':
                    self.current_link = value
                    self.in_link = True
                    self.link_text = ""
                    break
    
    def handle_endtag(self, tag):
        if tag.lower() == 'a' and self.in_link:
            if self.current_link and self.link_text.strip():
                # 只收集.html链接
                if self.current_link.lower().endswith('.html') or self.current_link.lower().endswith('.htm') or self.current_link.lower().endswith('.xhtml'):
                    self.links.append({
                        'href': self.current_link,
                        'text': self.link_text.strip()
                    })
            self.in_link = False
            self.current_link = None
            self.link_text = ""
    
    def handle_data(self, data):
        if self.in_link:
            self.link_text += data

class ImageExtractor(HTMLParser):
    """HTML图片提取器"""
    def __init__(self):
        super().__init__()
        self.images = []
    
    def handle_starttag(self, tag, attrs):
        if tag.lower() == 'img':
            src = None
            for name, value in attrs:
                if name.lower() == 'src':
                    src = value
                    break
            if src:
                self.images.append(src)

def extract_title_from_html(html_content):
    """从HTML内容中提取title标签内容"""
    extractor = TitleExtractor()
    try:
        extractor.feed(html_content)
        return extractor.title.strip() if extractor.title.strip() else None
    except:
        return None

def extract_links_from_html(html_content):
    """从HTML内容中提取所有链接"""
    extractor = LinkExtractor()
    try:
        extractor.feed(html_content)
        return extractor.links
    except Exception as e:
        print(f"  警告: 解析链接时出错: {e}")
        return []

def extract_images_from_html(html_content):
    """从HTML内容中提取所有图片链接"""
    extractor = ImageExtractor()
    try:
        extractor.feed(html_content)
        return extractor.images
    except Exception as e:
        print(f"  警告: 解析图片时出错: {e}")
        return []

def natural_sort_key(text):
    """自然排序键，正确处理数字"""
    def convert(text):
        return int(text) if text.isdigit() else text.lower()
    return [convert(c) for c in re.split('([0-9]+)', str(text))]

def fix_html_links(html_content, link_map):
    """修复HTML中的链接路径"""
    # 替换所有的href链接
    for old_href, new_href in link_map.items():
        # 匹配href="old_link"的模式
        pattern = rf'href=["\']({re.escape(old_href)})["\']'
        replacement = f'href="{new_href}"'
        html_content = re.sub(pattern, replacement, html_content, flags=re.IGNORECASE)
    
    return html_content
def fix_html_links_anchor(html_content, link_map):
    """修复HTML中的链接路径"""
    # 替换所有的href链接
    for old_href, new_href in link_map.items():
        # 匹配href="old_link"的模式
        pattern = rf'href=["\']({re.escape(old_href)})#'
        replacement = f'href="{new_href}#'
        html_content = re.sub(pattern, replacement, html_content, flags=re.IGNORECASE)
    
    return html_content

def fix_image_paths(html_content, image_map):
    """修复HTML中的图片路径"""
    for old_src, new_src in image_map.items():
        # 匹配src="old_path"的模式
        pattern = rf'src=["\']({re.escape(old_src)})["\']'
        replacement = f'src="{new_src}"'
        html_content = re.sub(pattern, replacement, html_content, flags=re.IGNORECASE)
    
    return html_content

def get_image_type(file_path):
    """根据文件扩展名获取图片MIME类型"""
    ext = file_path.lower().split('.')[-1]
    mime_types = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
        'bmp': 'image/bmp',
        'webp': 'image/webp',
        'svg': 'image/svg+xml'
    }
    return mime_types.get(ext, 'image/jpeg')

class BookVolume:
    """书卷类"""
    def __init__(self, path, volume_name):
        self.path = Path(path)
        self.volume_name = volume_name
        self.volume_number = self.extract_volume_number(volume_name)
        self.index_content = ""
        self.chapters = []
        self.images = []
        self.has_index = False
        self.chapter_link_map = {}  # 用于存储链接映射关系
        self.image_map = {}  # 用于存储图片映射关系
        self.all_headings = []  # 存储所有标题信息
        self.heading_extractor = HeadingExtractor()  # 标题提取器

    
    def extract_volume_number(self, volume_name):
        """从卷名中提取数字"""
        #match = re.search(r'第(\d+)卷', volume_name)
        match = re.search(r'(\d+)', volume_name)
        return int(match.group(1)) if match else 0
    
    def load_volume_index(self):
        """加载卷的总目录（index.html）"""
        #index_path = self.path / f"MEW{self.volume_number}.html"
        index_path = self.path / f"index.html"
        if index_path.exists():
            try:
                with open(index_path, 'r', encoding='utf-8') as f:
                    self.index_content = f.read()
                self.has_index = True
                #print(f"  加载卷目录: {self.volume_name}/index.html")
                # 从目录页提取图片
                self.collect_images_from_content(self.index_content, "index.html")
                
            except Exception as e:
                print(f"  警告: 无法读取 {index_path}: {e}")
        else:
            print(f"  注意: {self.volume_name} 没有找到 index.html")
    
    def collect_images_from_content(self, html_content, source_file):
        """从HTML内容中收集图片"""
        images = extract_images_from_html(html_content)
        for img_src in images:
            if not img_src.startswith(('http://', 'https://', 'data:')):
                # 相对路径图片
                img_path = self.path / img_src
                if img_path.exists():
                    img_id = f"ME{self.volume_number:02d}-img{len(self.images) + 1:03d}"
                    img_ext = img_path.suffix.lower()
                    img_filename = f"{img_id}{img_ext}"
                    
                    self.images.append({
                        'original_path': img_path,
                        'filename': img_filename,
                        'id': img_id,
                        'source_file': source_file,
                        'vol_num':self.volume_number
                    })
                    
                    # 建立映射关系
                    self.image_map[img_src] = img_filename
                    print(f"    收集图片: {img_src} -> {img_filename}")
    def rehead_ka(self,recontent):
        recontent=re.sub(r'''<(h[\d]) class="sub_chap3"[^<]*?>([\d]+?)\.\s+([\S\s]+?)</h[\d]>''',chapternumbergerman.kaptiel_no,recontent,flags=re.DOTALL|re.IGNORECASE)

        recontent=re.sub(r'''<(h[\d]) class="sub_chap1"[^<]*?>([IV]+?)\.\s+([\S\s]+?)</h[\d]>''',chapternumbergerman.abschnitt_no,recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r'''(<h[\d]) class="[\S]+?"[^<]*?>''',r'\1>',recontent,flags=re.DOTALL|re.IGNORECASE)
        return recontent
    def recont(self,recontent):
        def footnote(match):
            footnotes=match.group(1)
            footnotes=re.sub(r"<p>([A]*[\d]{1,2})\s+([\S])",r"""<p class="fni"><a href="#fn\1ref" id="fn\1">\1</a> \2""",footnotes,flags=re.DOTALL|re.IGNORECASE)
            footnotes=r"""<aside class="fn">
<div class="fnt">Fußnoten</div>
"""+footnotes+r"</aside>"
            return footnotes
        recontent=re.sub(r"""<a class="page" href="http[\S]+?">\[([\d]+?)\]</a>""",r'<a id="S\1"></a>',recontent,flags=re.DOTALL|re.IGNORECASE)
        #recontent=re.sub(r"[\r\n\s]{2,}",r" ",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml" lang="de" xml:lang="de">[\s\r\n\S]+?<div class="zenoCOMain">""","<body>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<div class="zenoTRNavBottom">[\s\r\n\S]+?</html>""",r"</body>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<div class="zenoCOFooter">[\s\r\n\S]+?</html>""",r"</body>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<div class="zenoTXFnTable">\s*<table>""",r"""<aside class="fn">
<div class="fnt">Fußnoten</div>
""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""</table>\s*</div>\s*</body>""",r"</aside>\n</body>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<tr>\s+<td>""",r"<tr><td>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""</td>\s+</tr>""",r"</td></tr>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r""" pp="no">""",r">",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r'<p><a (?:id|name)=([\S]+?)(?: class="[\S ]+?")*></a><a(?: class="[\S ]+?")* href="([\S]+?)"(?: class="[\S ]+?")*>([\S]+?)</a> ',r'<p class="fni"><a href=\2 id=\1>\3</a> ',recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<a href=[\S]+? class="zenoTXKonk" title="Vorlage" name="([\d]+?)">\[[\d]+?\]</a>""",r"<a id=S\1></a>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<a href=[\S]+? name=("[\S]+?") class="zenoTXFnRef">([A-Za-z\d]+?)</a>""",r"""<sup><a id=\1 href="#fn\2">\2</a></sup>""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r'<a(?: class="[^<]+?")* (?:id|name)=("[\S]+?")(?: class="[^<]+?")*>\s*</a>\s*<a(?: class="[^<]+?")* href=([\S]+?)(?: class="[^<]+?")*>(?:<sup>)*([\S]+?)(?:</sup>)*</a>',r'<sup><a href=\2 id=\1>\3</a></sup>',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r"""<tr><td><a\s*href=[\S]+?\s*name=("[\S]+?")\s*class="zenoTXFnText">([A\d]+?)</a></td>\s*<td><p>([\S\s\r\n]+?)</p></td></tr>""",r"""<p class="fni"><a id=\1 href="#fn\2ref">\2</a> \3</p>""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<a\s*href=[\S]+?\s*name=("[\S]+?")\s*class="zenoTXFnText">([A\d]+?)</a>""",r"""<a id=\1 href="#fn\2ref">\2</a>""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r'<p class="zenoPR">([\s\r\n\S]+?)</p>',r'<p class="rgt">\1</p>',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r"""<p class="zenoPLm8n12"><span class="zenoTXColor2">([\S\s\r\n]+?)</span>(<a [\S ]+?</a>)*\s*</p>""",r'<blockquote><p class="poem">\1\2</p></blockquote>',recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""</p></blockquote>\s*<blockquote><p class="poem">""",'</p>\n<p class="poem">',recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r'<p class="zenoPLm8n12">',r'<p class="poem">',recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<p class="zenoPLm4n0"><span class="zenoTXColor2">([\S\s\r\n]+?)</span>(<a [\S ]+?</a>)*\s*</p>""",r"<blockquote>\1\2</blockquote>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<p class="zenoPC">""",r"""<p class="ctr">""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<p class="zenoPLm4n0">""",r"<p>",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<p id="an[\d]+?"\s*/>""",r"",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<h[\d]>(Fußnote[n]*|Lesart[en]*)</h[\d]>([\S\s\r\n]+?)(</body>|<h[\d])""",r"""<aside class="fn">
<div class="fnt">\1</div>
\2
</aside>
\3""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<p class="fn">Fußnoten</p>\s*<div class="emptyLine">&nbsp;</div>([\s\r\n\S]+?)<div class="emptyLine">&nbsp;</div>""",footnote,recontent,flags=re.DOTALL|re.IGNORECASE)

        recontent=re.sub(r"""<sup>([\d]+?)</sup>(?!/|/<sub>)""",r"""<sup><a href="#fn\1" id="fn\1ref">\1</a></sup>""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<sup>([A\d]+?)</sup>(?!/|/<sub>)""",r"""<sup><a href="#fn\1" id="fn\1ref">\1</a></sup>""",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r"""<div class="emptyLine">&nbsp;</div>""",r"",recontent,flags=re.DOTALL|re.IGNORECASE)
        recontent=re.sub(r'([a-zA-Z])[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]*([\.,;])',r'\1\2\3',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r'([a-zA-Z])[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]*<',r'\1\2<',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r'[\.]+[ ]*(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]+',r'\1. ',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r'[\.]+[ ]*(<sup><a[^<]+?>[\S]+?</a></sup>)',r'\1.',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r'>[ ]*(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]+([\.,;])',r'>\1\2',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r'[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)</h',r'>\1</h',recontent ,flags=re.DOTALL | re.IGNORECASE)
        recontent=re.sub(r'(</(?:p|blockquote)>)[\s\r\n]*(<sup><a[^<]+?>[\S]+?</a></sup>)',r'\2\1',recontent ,flags=re.DOTALL | re.IGNORECASE)
        #recontent=re.sub(r"</p>",r"</p>\n",recontent,flags=re.DOTALL|re.IGNORECASE)
        #recontent=self.rehead(recontent)
        return recontent
    def contentre(self,recontent):
        content=re.sub(r'<html lang="de">[\s\r\n\S]+?<div class="nav top">[\s\r\n\S]+?</a>\s*</div>',r'<body>',recontent ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<body class="calibre">',r'<body>',content ,flags=re.DOTALL | re.IGNORECASE)
        
        content=re.sub(r'<p class="calibre12"><span>([\s\r\n\S]+?)</span></p>',r'<blockquote>\1</blockquote>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<p class="calibre19">([\s\r\n\S]+?)</p>',r'<blockquote>\1</blockquote>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<p class="calibre(?:10|64)">([\s\r\n\S]+?)</p>',r'<p class="ctr">\1</p>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<p class="calibre(?:18|25|60)">([\s\r\n\S]+?)</p>',r'<p class="rgt">\1</p>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r' class="calibre_[\S]*?"','',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r' class="calibre[\d]*?"','',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<div>[\s\r\n]+<span class="sgc2"><a href="[\S]+?#inhalt" class="pcalibre pcalibre1">Inhaltsverzeichnis</a></span>[\s\r\n]+</div>','',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<div>[\s\r\n]+<span><a href="[\S]+?#inhalt" class="pcalibre pcalibre1">Inhaltsverzeichnis</a></span>[\s\r\n]+</div>','',content ,flags=re.DOTALL | re.IGNORECASE)
        content = re.sub(r'''href="\.\./Text/''',r'href="',content ,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'(?:<(?:sup|span class="[^<]+?")>)*<a(?: class="[^<]+?")* href=([\S]+?) id=([\S]+?)(?: class="[^<]+?")*>([\S]+?)</a>(?:</(?:sup|span)>)*',r'<sup><a href=\1 id=\2>\3</a></sup>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'(?:<(?:sup|span class="[^<]+?")>)*<a(?: class="[^<]+?")* id=([\S]+?) href=([\S]+?)(?: class="[^<]+?")*>([\S]+?)</a>(?:</(?:sup|span)>)*',r'<sup><a href=\2 id=\1>\3</a></sup>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<a(?: class="[^<]+?")* href=([\S]+?) id=([\S]+?)(?: class="[^<]+?")*>(?:<(?:sup|span class="[^<]+?")>)*([\S]+?)(?:</(?:sup|span)>)*</a>',r'<sup><a href=\1 id=\2>\3</a></sup>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<a(?: class="[^<]+?")* id=([\S]+?) href=([\S]+?)(?: class="[^<]+?")*>(?:<(?:sup|span class="[^<]+?")>)*([\S]+?)(?:</(?:sup|span)>)*</a>',r'<sup><a href=\2 id=\1>\3</a></sup>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'[\[]*<a\s+name=([\S]+?)>\s*</a>\s*<a\s+href=([\S]+?)>[\s\[]*([\d]+?)[\s\]]*</a>[\]]*',r'<sup><a href=\2 id=\1>\3</a></sup>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<a(?: class="[^<]+?")* id=([\S]+?)(?: class="[^<]+?")*>\s*</a>\s*<a(?: class="[^<]+?")* href=([\S]+?)(?: class="[^<]+?")*><sup[^<]*?>([\S]+?)</sup></a>',r'<sup><a href=\2 id=\1>\3</a></sup>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<a(?: class="[^<]+?")* id=("[\S]+?")(?: class="[^<]+?")*>\s*</a>\s*<a(?: class="[^<]+?")* href=([\S]+?)(?: class="[^<]+?")*>([\S]+?)</a>',r'<sup><a href=\2 id=\1>\3</a></sup>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'<p class="ind1" id=([\S]+?)>(?:<sup>)*<a href=([\S]+?)(?: class="[^<]+?")*>([\S]+?)</a>(?:</sup>)* ',r'<p class="fni"><a href=\2 id=\1>\3</a> ',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<p class="ind1">(?:<sup>)*<a(?: class="[^<]+?")* id=([\S]+?) href=([\S]+?)(?: class="[^<]+?")*>([\S]+?)</a>(?:</sup>)* ',r'<p class="fni"><a href=\2 id=\1>\3</a> ',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<p class="ind1">(?:<sup>)*<a(?: class="[^<]+?")* href=([\S]+?) id=([\S]+?)(?: class="[^<]+?")*>([\S]+?)</a>(?:</sup>)* ',r'<p class="fni"><a href=\1 id=\2>\3</a> ',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<p class="ind1"><a id=([\S]+?) class="[\S ]+?"></a><a class="[\S ]+?" href=([\S]+?)>([\S]+?)</a> ',r'<p class="fni"><a href=\2 id=\1>\3</a> ',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<p class="ind1"><a id=([\S]+?) class="[\S ]+?"></a><a class="[\S ]+?" href="(#[\S]+?)">([\S]+?)</a> ',r'<p class="fni"><a href=\2 id=\1>\3</a> ',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<p class="ind">',r'<p>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<p class="ind[\d]">',r'<p>',content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r'<a name=',r'<a id=',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'[-]*[ ]{0,2}<a id=("S[\d]+?")></a>',r'<a id=\1></a>',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'([a-zA-Z])[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]*([a-zA-Z&])',r'\1\2 \3',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'[-]*\s+(<a id=["]*S[\S]+?["]*></a>)[ ]+',r'\1 ',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'[-]+\s+(<a id=["]*S[\S]+?["]*></a>)([a-zA-Z])',r'\1\2',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'([a-zA-Z])[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]*([\.,;])',r'\1\2\3',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'([a-zA-Z])[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]*<',r'\1\2<',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'[\.]+[ ]*(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]+',r'\1. ',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'[\.]+[ ]*(<sup><a[^<]+?>[\S]+?</a></sup>)',r'\1.',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'>[ ]*(<sup><a[^<]+?>[\S]+?</a></sup>)[ ]+([\.,;])',r'>\1\2',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'[ ]+(<sup><a[^<]+?>[\S]+?</a></sup>)</h',r'>\1</h',content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'(</(?:p|blockquote)>)[\s\r\n]*(<sup><a[^<]+?>[\S]+?</a></sup>)',r'\2\1',content ,flags=re.DOTALL | re.IGNORECASE)
        content = re.sub(r"<([/]*)em>",r"<\1i>",content ,flags=re.DOTALL | re.IGNORECASE)
        content=re.sub(r'(<h[\d][^<]*?>)<b>([\S ]+?)</b>(</h[\d]>)',r'\1\2\3',content,flags=re.DOTALL|re.IGNORECASE)
        content = re.sub(r"""<span class="italic">([\S ]+?)</span>""",r"<i>\1</i>",content ,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r"""<div class="mbp_pagebreak" id="calibre_pb_[\d]+?"></div>""",'',content ,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r"""<h[\d][^<]*?>(Fußnote[n]*|Lesart[en]*)</h[\d]>([\S\s\r\n]+?)(</body>|<h[\d])""",r"""<aside class="fn">
<div class="fnt">\1</div>
\2
</aside>
\3""",content,flags=re.DOTALL|re.IGNORECASE)
        content=re.sub(r"""<body>((?:(?!<aside)[\s\r\n\S]+?))<p class="fni">([\s\r\n\S]+?)</body>""",r'''<body>\1
<aside class="fn">
<p class="fni">\2</aside>
</body>''',content ,flags=re.DOTALL|re.IGNORECASE)
        content = content.replace(r"&szlig;", r"ß")
        content = content.replace(r"&ouml;", r"ö")
        content = content.replace(r"&auml;", r"ä")
        content = content.replace(r"&uuml;", r"ü")
        content = content.replace(r"&Szlig;", r"ẞ")
        content = content.replace(r"&Ouml;", r"Ö")
        content = content.replace(r"&Auml;", r"Ä")
        content = content.replace(r"&Uuml;", r"Ü")
        if self.volume_number in range(23,26):
            content=self.rehead_ka(content)
        return content
    
    def scan_chapters(self,volfilename=[]):
        """通过解析目录页链接来扫描章节文件"""
        if not self.has_index:
            print(f"  {self.volume_name} 没有目录页，跳过")
            return
        
        # 从目录页提取所有链接
        links = extract_links_from_html(self.index_content)
        
        if not links:
            print(f"  {self.volume_name} 目录页中没有找到链接")
            return
        
        print(f"  从目录页找到 {len(links)} 个链接:")
        
        chapter_number = 1
        for link in links:
            href = link['href']
            link_text = link['text']
            
            # 构建完整的文件路径
            article_path = self.path / href
            
            if not article_path.exists():
                print(f"    跳过: {href} (文件不存在)")
                continue
            if len(self.chapters) > 0 and self.chapters[-1]['original_file']==href:
                continue
            try:
                try:          
                    encodings = ['utf-8-sig','utf-8','iso-8859-1',  'gbk', 'cp1252']
                    for encoding in encodings:
                        with open(article_path, 'r', encoding=encoding, errors='ignore') as f:
                            content = f.read()
                            break  # 成功读取后跳出循环
                except Exception as e:
                    raise Exception(f"所有解码尝试都失败: {e}") 

                    # 如果所有编码都失败，尝试忽略错误
                
    


                content=content.replace("<br>&emsp;&emsp;<br>&emsp;&emsp;","<br>&emsp;&emsp;")
                content=re.sub(r"<br>&emsp;&emsp;[\s]*</h1>",r"</h1>",content,flags=re.DOTALL|re.IGNORECASE)
                content=content.replace('</p><p class="quote">','')
                content=re.sub(r'<html lang="de"></html>','',content,flags=re.DOTALL|re.IGNORECASE)
                content=re.sub(r'','',content,flags=re.DOTALL|re.IGNORECASE)
                content=re.sub(r'<!--(&#[\S]+?;)-->',r'\1', content, flags=re.DOTALL | re.IGNORECASE)
                content=re.sub(r'<!--[\S\r\n\s]+?-->',r'', content, flags=re.DOTALL | re.IGNORECASE)
                title = extract_title_from_html(content)
                if not title or title == "":
                    title = link_text if link_text else f"第{chapter_number}章"
                title=title.replace("<","&lt;")
                title=title.replace(">","&gt;")
                if href.startswith(("2000")):
                    content=self.recont(content) 
                else:
                    content=self.contentre(content)
                if volfilename and len(volfilename) >= chapter_number and volfilename[chapter_number-1] and isinstance(volfilename[chapter_number-1], int):
                    chapter_filename = f"ME{self.volume_number:02d}-{volfilename[chapter_number-1]:03d}.html"
                elif volfilename and len(volfilename) >= chapter_number and volfilename[chapter_number-1] and volfilename[chapter_number-1].startswith(("ME")):
                    chapter_filename = volfilename[chapter_number-1] + ".html"
                else:
                    chapter_filename = href.replace(".xhtml",".html")
                #
                # 收集文章中的图片
                content= content.replace("<sup>", "￥￥￥<sup>")
                content= content.replace("</sup>", "</sup>￥￥￥") 
                if '-FN' in chapter_filename:
                    content=content.replace('<aside class="fn">',"").replace("</aside>","")

                self.collect_images_from_content(content, href)

                # 提取标题并修改HTML（添加ID）
                modified_content, headings = self.heading_extractor.extract_headings_from_html(
                    content, href, chapter_number,title,self.volume_number
                )
                
                # 存储标题信息
                self.all_headings.extend(headings)
                
                #if headings:
                    #print(f"    第{chapter_number}章: 找到 {len(headings)} 个标题")
                
                # 尝试从HTML的title标签提取标题，如果失败则使用链接文本

                #chapter_filename=chapter_filename.replace("_","-")
                #chapter_filename=chapter_filename.replace("me","ME")
                #chapter_filename = f"ME{self.volume_number}-{chapter_number:03d}.html"
                if headings and self.volume_number!=3:
                    title=headings[0]['text']
                
                self.chapters.append({
                    'number': chapter_number,
                    'title':title ,
                    'content': modified_content,  # 使用修改后的内容（包含标题ID）
                    'filename': chapter_filename,
                    'original_file': href,
                    'link_text': link_text,
                    'headings': headings  # 存储本章的标题信息
                })
                
                # 建立链接映射关系
                self.chapter_link_map[href] = chapter_filename
                
                chapter_number += 1
                
            except Exception as e:
                print(f"    警告: 无法读取文章文件 {href}: {e}")
                continue

class EpubBookBuilder:
    """EPUB图书构建器"""
    def __init__(self, title, author="Unknown", language="de",ws=[],ws2=[]):
        self.title = title
        self.author = author
        self.language = language
        self.uuid = str(uuid.uuid4())
        self.volumes = []
        self.all_images = []
        self.css = r"""<!--
.style1 {color: #FF0000;font-family: "黑体";}
.style2 {font-size: 1.25em}
.style3 {font-size: 0.75em;text-align:center;}
.quote {font-size: 0.75em;margin: 1.5em 1px 1.5em 1px;}
b {font-family: "黑体,Times New Roman,sans serif";}
table {margin: 1.5em;max-width:100%;}
table.quote  {font-size: 0.75em; margin: 1.5em;}
h1,h2 {color: #FF0000; font-family: "黑体"; text-align:center;}
body {background-color: #D1E3FE;}
span.cq,div.cq {text-align:center;display: block;}
div[align="right"] table {margin: 1.5em 0 1.5em auto;}
span[align="center"] {display: block;}
-->
"""
        self.ws=ws
        self.volfilename=ws2
    def scan_book_structure(self, book_dir):
        """扫描整本书的结构"""
        book_path = Path(book_dir)
        if not book_path.exists():
            print(f"错误: 书籍目录 '{book_dir}' 不存在")
            return
        
        print(f"扫描书籍目录: {book_dir}")
        
        # 查找所有"第n卷"目录
        volume_dirs = []
        for item in book_path.iterdir():
            #if item.is_dir() and re.match(r'第\d+卷', item.name):
            if item.is_dir() and re.match(r'\d+', item.name):
            #if item.is_dir() and re.match(r'me25', item.name):
                volume_dirs.append(item)
        
        if not volume_dirs:
            print("错误: 没有找到符合'第n卷'格式的目录")
            return
        
        # 按卷号排序
        volume_dirs.sort(key=lambda x: natural_sort_key(x.name))
        #ctl=[]
        #ctl=[20,20]
        #invo=1        
        # 处理每一卷
        for vol_dir in volume_dirs:
            #if ctl:
            #    if invo not in range(ctl[0],ctl[1]+1):
            #        invo+=1
            #        continue
            #    invo+=1
            print(f"\n处理卷: {vol_dir.name}")
            volume = BookVolume(vol_dir, vol_dir.name)
            volume.load_volume_index()
            
            
            if volume.has_index:
                i=volume.volume_number
                for row in self.ws.iter_rows(min_row=i, max_row=i, values_only=True):
                    volume_info=row[1]
                for col in self.volfilename.iter_cols(min_col=i, max_col=i, values_only=True):
                    column_data = list(col)  # col 就是整个列的所有值
                volume_filenames=column_data
                if volume_filenames:
                    volume.scan_chapters(volume_filenames)
                else:
                    volume.scan_chapters()
                volume.volume_name=volume_info
                self.volumes.append(volume)
                self.all_images.extend(volume.images)
                total_headings = sum(len(chapter['headings']) for chapter in volume.chapters)

                #print(f"  完成: 找到 {len(volume.chapters)} 个章节, {len(volume.images)} 张图片, {total_headings} 个标题")
            else:
                print(f"  警告: {vol_dir.name} 中没有找到有效章节")
    
    def create_volume_index_toc(self):
        index_content=''
        for volume in self.volumes:
            i=volume.volume_number
            volumetext=volume.volume_name
            index_content += f'''<a href="{volume.volume_number}/index.html" target=_blank>{volumetext}</a><br>\n'''
        return index_content
            
    def create_volume_index_with_toc(self, volume):
        """创建包含详细目录的卷索引页"""
        i=volume.volume_number
        volumetext=f"{volume.volume_name}"
        selftitle=self.title
        ish4=""
        subti=""
        BAND=r"BAND "
        #if i==5:
        #    BAND+=f"{i}-{i+1}"
        #elif i==10:
        #    BAND+=f"{i-2}-{i}"
        #elif i==18:
        #    BAND+=f"{i}-{i+2}"
        BAND+=f"{i}"
        #if i==7:
        #    subti="\n<h4>(Naturrecht und Staatswissenschaft im Grundrisse. Zum Gebrauch für seine Vorlesungen)</h4>"
        #    ish4=",h4"
        volumetext=f"<h3>{volumetext}</h3>"
        if i==20:
            volumetext=f"<h3>Anti-Dühring<br>Dialektik der Natur</h3>\n<h4>（反杜林论、自然辩证法）</h4>"
            ish4=",h4"
        if i==26:
            volumetext=f"<h3>Theorien über den Mehrwert, Zweiter Teil</h3>\n<h4>（剩余价值理论，第二卷）</h4>"
            ish4=",h4"
        if i==40:
            volumetext=f"<h3>Karl Marx: Schriften und Briefe, November 1837 bis August 1844</h3>\n<h4>（马克思早期著作）</h4>"
            ish4=",h4"
        if i==42:
            volumetext=f"<h3>Karl Marx<br>Ökonomische Manuskripte 1857/1858</h3>\n<h4>（马克思1857-1858年经济学手稿）</h4>"
            ish4=",h4"    
        if i==50:
            BAND="Weitere Werke"   
        index_content = f'''<html lang="de">
<head>
<title>{selftitle} {BAND} – {volume.volume_name}</title>'''
        index_content+=r'''<META content="text/html; charset=utf-8" http-equiv="Content-Type"/>
<META name="viewport" content="width=device-width, initial-scale=1.0"/>
<link rel="stylesheet" type="text/css" href="../mewde.css"/>
<script src="/mlr.js"></script>
<style type="text/css">
<!--
h2,h3 {font-family:"Times New Roman";margin:0;}
--></style>
</head>
<body>
'''
        if ish4:
            index_content=index_content.replace(r"h2,h3 ",f"h2,h3{ish4} ")
        index_content_title=f'''<h2 style="color: #DC3545;">{selftitle}<br>BAND {i}</h2>
{volumetext}'''
        if i==50:
            index_content_title=f'''<h2 style="color: #DC3545;">{selftitle}</h2>
<h3>Weitere Werke</h3>'''
        index_content+=index_content_title+'''
<h2>Inhalt</h2>
<nav class="TOC">
<ol>'''
        json_list=[]
        def sub_navpoint(headings):
            """使用 while 循环构建嵌套导航结构"""
            result = ""
            i=0
            while i<len(headings):
                heading=headings[i]
                filename=heading['filename']
                if heading["tag"]=="title":
                    resultplus= f'<li><a href="{filename}" target=_blank>{heading["text"]}</a>'
                else:
                    resultplus= f'<li><a href="{filename}#{heading["id"]}" target=_blank>{heading["text"]}</a>'

                if heading["text"] in [r"人名索引",r"文学作品和神话中的人物索引",r"文献索引",r"报刊索引"]:
                    result+=resultplus+'</li>\n'
                    break
                result+=resultplus
                sub_headings = []
                next_index = headings.index(heading) + 1
                j=next_index
                while j < len(headings) and headings[j]['level'] > heading['level']:
                    sub_headings.append(headings[j])
                    j += 1
                if sub_headings:
                    result += '<ol>\n' + sub_navpoint(sub_headings) + '</ol>\n'
                    i+=len(sub_headings)+1
                else:
                    i+=1
                result += '</li>\n'
            return result
        # 添加章节和标题的目录
        headings=[]
        volIII=0
        for chapter in volume.chapters:
            headings_final=[]
            chapter_path=f"{chapter['filename']}"
            # 添加章节内的标题子目录            
            if chapter['headings']:
                for heading in chapter['headings']:
                    text = heading['text']
                    level = heading['level']
                    headid = heading['id']
                    tag=heading['tag']
                    if headid.endswith("-1"):
                        tag="title"
                    headings_final.append({
                        'tag':tag,
                        'text':text,
                        'level':level,
                        'id':headid,
                        'source_file':chapter_path,
                        'chapter_number':chapter['number'],
                        'filename':chapter['filename']
                        })
                    #if chapter['filename'].startswith("ME22"):
                    #    print(f'<li><a href="{volume.volume_number}/{chapter["filename"]}" target=_blank>{heading["text"]}</a></li>')
            json_list.append({
            "file":     chapter['filename'],
            "title":    chapter['title'],
            "headings":headings_final
            })
            headings.extend(headings_final)
        index_content += '\n'+sub_navpoint(headings)  

        if volume.volume_number==22:
            index_content += r'''
</ol>
<h3>Vorworte und Einleitungen</h3>
<ol>
<li><a href="../4/ME22-052.html" target=_blank>Vorwort zur vierten deutschen Ausgabe (1890) des "Manifests der Kommunistischen Partei"</a></li>
<li><a href="../17/ME22-188.html" target=_blank>Einleitung zu Karl Marx' "Der Bürgerkrieg in Frankreich" (Ausgabe 1891)</a></li>
<li><a href="../4/ME22-200.html" target=_blank>Friedrich Engels - Zur spanischen Ausgabe von Karl Marx' "Elend der Philosophie"</a></li>
<li><a href="../6/ME22-202.html" target=_blank>Einleitung zu Karl Marx' "Lohnarbeit und Kapital" (Ausgabe 1891)</a></li>
<li><a href="../19/ME22-210.html" target=_blank>Vorwort zur vierten Auflage (1891) der "Entwicklung des Sozialismus von der Utopie zur Wissenschaft"</a></li>
<li><a href="../21/ME22-211.html" target=_blank>Vorwort zur vierten Auflage des "Ursprungs der Familie, des Privateigentums und des Staats</a></li>
<li><a href="../2/ME22-265.html" target=_blank>Vorwort zur englischen Ausgabe (1892) der "Lage der arbeitenden Klasse in England"</a></li>
<li><a href="../4/ME22-282.html" target=_blank>Vorwort zur zweiten polnischen Ausgabe (1892) des "Manifests der Kommunistischen Partei"</a></li>
<li><a href="../19/ME22-287.html" target=_blank>Einleitung [zur englischen Ausgabe (1892) der "Entwicklung des Sozialismus von der Utopie zur Wissenschaft"]</a></li>
<li><a href="../4/ME22-365.html" target=_blank>An den italienischen Leser - Vorwort zur italienischen Ausgabe (1893) des "Manifests der Kommunistischen Partei"</a></li>
<li><a href="../18/ME22-419.html" target=_blank>Vorbemerkung (1894) [zu "Die Bakunisten an der Arbeit. Denkschrift über den Aufstand in Spanien im Sommer 1873"]</a></li>
<li><a href="../18/ME22-421.html" target=_blank>Nachwort (1894) [zu "Soziales aus Rußland"]</a></li>
<li><a href="../7/ME22-509.html" target=_blank>Einleitung zu Karl Marx' "Klassenkämpfe in Frankreich 1848 bis 1850" (1895)</a></li>
'''           
        index_content += '''</ol></nav>
</body>
</html>'''
        
        return index_content,json_list

        


    def build_epub_folder(self, output_folders):
        """构建EPUB文件夹结构"""
        if not self.volumes:
            print("错误: 没有找到任何卷，无法生成EPUB")
            return
        outpath=[]

        for output_folder in output_folders:

            output_path = Path(output_folder)
            #if output_path.exists():
            #    shutil.rmtree(output_path)
            output_path.mkdir(exist_ok=True)
            outpath.append(output_path)
            print(f"\n构建EPUB文件夹: {output_folder}")        

        
    
        '''with open(output_path / "index.html", 'w', encoding='utf-8-sig', newline='') as f:
            f.write(self.create_contents_html())'''

        # 创建内容文件
        total_chapters = 0
        total_headings = 0
        
        for volume in self.volumes:
            # 处理卷目录文件（包含详细目录）
            volpath=[]
            for output_path in outpath:
                vol_output_Path=output_path/ Path(f'{volume.volume_number}')
                vol_output_Path.mkdir(exist_ok=True)
                volpath.append(vol_output_Path)
            k=volume.volume_number

           
            #if volume.has_index and volume.volume_number!=50:
            if volume.has_index:
                vol_index_filename = f"index.html"
                index_content,json_list = self.create_volume_index_with_toc(volume)
                manifest_json = json.dumps(json_list, ensure_ascii=False, indent=None)
                for vol_output_Path in volpath:
                    with open(vol_output_Path/vol_index_filename, 'w', encoding='utf-8-sig', newline='\r\n') as f:
                        f.write(index_content)
                    with open(vol_output_Path/f"index.json", 'w', encoding='utf-8') as f:
                        f.write(manifest_json)
                '''vol_index_filename = f"index-{volume.volume_number}.html"
                index_content = self.create_vol_index(volume)
                with open(vol_output_Path/vol_index_filename, 'w', encoding='utf-8-sig', newline='') as f:
                    f.write(index_content)'''
                
            
            
            # 处理章节文件
            for chapter in volume.chapters:
                # 修复章节中的图片路径，删除原有CSS，并添加新CSS链接
                fixed_content = fix_image_paths(chapter['content'], volume.image_map)
                fixed_content = fix_html_links_anchor(chapter['content'], volume.chapter_link_map)
                

                fixed_content =fixed_content.replace('<br/>','<br>')
                #fixed_content=fixed_content.replace('</aside>','</p>')
                fixed_content=fixed_content.replace('</p><p class="quote">','')
                fixed_content=fixed_content.replace('class="date"','class="style3"')
                fixed_content=fixed_content.replace('］',']')
                fixed_content=fixed_content.replace('［','[')
                #fixed_content = re.sub(r'<style[^>]*>.*?</style>', f'''<style type="text/css">{self.css}</style>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)

                #fixed_content = re.sub(r'<link[^>]*rel=["\']stylesheet["\'][^>]*>', f'''<style type="text/css">{self.css}</style>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content = re.sub(r'<div class=["\']style3["\']>', '<div class="date">', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                
                
                #fixed_content=re.sub(r'(<br>[\s\r\n ]*){2,}<br>','<br>  ',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                #fixed_content=re.sub(r'(<br>[\s\r\n ]*){2,}','<br>\n',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                fixed_content=fixed_content.replace(r'<br>\s*(<a id=[\S]></a>)\s*<br>',f'<br>\n'+r'\1')
                fixed_content=fixed_content.replace(r'<br>  <a href=',f'<br>\n<a href=')
                fixed_content=re.sub(r'(<p>)*Textvarianten</p>','''
<span style="font-size:1.2em;"><i>Textvarianten</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'<P><HR/></P>[\s\r\n]*<P>(Anmerkungen[ \S]+?)</P>',r'''<P><HR/></P><p style="font-size:1.2em;"><i>\1</i></p><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                if volume.volume_number in range(21,23):
                    fixed_content=re.sub(r'(<p>)*Fußnoten von Friedrich Engels</p>','''
<span style="font-size:1.2em;"><i>Fußnoten</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                    fixed_content=re.sub(r'(<p>)*Fußnoten von Engels</p>','''
<span style="font-size:1.2em;"><i>Fußnoten</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                else:
                    fixed_content=re.sub(r'(<p>)*Fußnoten von Friedrich Engels</p>','''
<span style="font-size:1.2em;"><i>Fußnoten von Engels</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                    fixed_content=re.sub(r'(<p>)*Fußnoten von Engels</p>','''
<span style="font-size:1.2em;"><i>Fußnoten von Engels</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'(<p>)*Fußnoten von Marx und Engels</p>','''
<span style="font-size:1.2em;"><i>Fußnoten von Marx und Engels</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'(<p>)*Fußnoten von Marx</p>','''
<span style="font-size:1.2em;"><i>Fußnoten von Marx</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'(<p>)*Fußnoten von Karl Marx</p>','''
<span style="font-size:1.2em;"><i>Fußnoten von Marx</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'(<p>)Fußnoten</p>','''
<span style="font-size:1.2em;"><i>Fußnoten</i></span><BR><BR>''', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'</p>\s*</body>','</p>\n</body>', fixed_content, flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'</p>\s*<hr/>\s*</body>',r'</p></body>',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'</p>\s*<hr/>\s*</p>\s*\s*</body>',r'</p></body>',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'￥￥￥',r'',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                fixed_content=fixed_content.replace(r'',f'')
                #fixed_content=re.sub(r'<br>([\s\r\n\S]+?)<br>',r'<p>\1</p>',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'<br></p>',r'</p>',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                fixed_content=re.sub(r'<p></p>',r'',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                #if k==42:
                #    fixed_content=re.sub(r'href="\.\./mewde.css',r'href="../mew42.css',fixed_content ,flags=re.DOTALL | re.IGNORECASE)
                wfilename=chapter['filename']
                if chapter['filename'] in (r"ME23-000.html",r"ME24-000.html",r"ME25-000.html"): #
                    wfilename=f"index.html"
                for vol_output_Path in volpath:
                    with open(vol_output_Path/wfilename, 'w', encoding='utf-8-sig', newline='\r\n') as f:
                        f.write(fixed_content)
                
                total_chapters += 1
                total_headings += len(chapter['headings'])
        # 复制图片文件
        for img in self.all_images:
            try:
                for output_path in outpath: 
                    dest_path = output_path/ f'{img['vol_num']}'/img['filename']
                    shutil.copy2(img['original_path'], dest_path)
                print(f"  复制图片: {img['filename']}")
            except Exception as e:
                print(f"  警告: 复制图片失败 {img['original_path']}: {e}")
        
        print(f"\nEPUB构建完成!")
        print(f"- 总卷数: {len(self.volumes)}")
        print(f"- 总章节数: {total_chapters}")
        print(f"- 总标题数: {total_headings}")
        print(f"- 总图片数: {len(self.all_images)}")
        print(f"- 输出目录: {output_folder}")
        
        # 显示详细的统计信息
        print(f"\n详细统计:")
        for volume in self.volumes:
            vol_headings = sum(len(chapter['headings']) for chapter in volume.chapters)
            print(f"  {volume.volume_name}: {len(volume.chapters)} 章节, {vol_headings} 标题")
        
        print(f"\n导航特性:")
        print(f"- 支持多级标题导航 (h1-h6)")
        print(f"- 标题已添加唯一ID用于锚点链接")
        print(f"- NCX文件包含完整的层级导航结构")
        print(f"- 卷目录页包含详细的标题索引")
def main():
    """主程序"""
    print("=== 完整书籍到EPUB转换器 ===\n")
    
    # 配置参数
    book_title = "KARL MARX FRIEDRICH ENGELS WERKE"  # 修改为你的书名
    book_author = "Georg Wilhelm Friedrich Hegel"      # 修改为作者名
    book_dir = r"./MEADE"  # 修改为你的书籍根目录
    #output_dir = "./MEA-DE"  # 输出目录名
    output_dirs = [r"./mlread/docs/MEW-ZENO",r"./MARX-ZH-CN.github.io1/docs/MEW-ZENO",r"./MARX-ZH-CN-node/docs/MEW-ZENO"] 
    #output_dirs=[r"./MEA-DE1/"]

    excel_file=Path(r"LENIN-toc.xlsx")
    # 创建构建器
    wb = openpyxl.load_workbook(excel_file)            
    sheet = wb.active
    ws = wb['Sheet1']    
    wb2 = openpyxl.load_workbook(Path(r"MEW-part.xlsx"))
    ws2= wb2['Sheet1']   
    # 创建构建器
    builder = EpubBookBuilder(book_title, book_author, "zh",ws,ws2)

    # 扫描书籍结构
    builder.scan_book_structure(book_dir)
    
    if builder.volumes:
        # 构建EPUB
        builder.build_epub_folder(output_dirs)
        
        print(f"\n=== 处理完成 ===")
        print("EPUB文件夹已创建，现在可以正常显示目录并跳转了！")
        print("\n压缩为EPUB文件的方法:")
        print("1. 选择文件夹内的所有文件和文件夹")
        print("2. 压缩为ZIP文件")
        print("3. 重命名为.epub扩展名")
        print("\n注意: 确保目录页的链接现在可以正确跳转到对应文章")
    else:
        print("\n错误: 没有找到有效的书籍结构")
        print("请确保:")
        print("1. 有'第n卷'格式的文件夹")
        print("2. 每卷中有index.html和对应的文章文件")
        print("3. HTML文件包含有效的title标签")

if __name__ == "__main__":
    main()

